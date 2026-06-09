import asyncio
import aiohttp
import gspread_asyncio
from oauth2client.service_account import ServiceAccountCredentials
from bs4 import BeautifulSoup
import logging
import re
import random
import time
import json
from openpyxl.utils import get_column_letter
from typing import Dict, List, Tuple, Set, Optional, Any

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('/home/tracxn-ds-528/Desktop/OD_OR_Publishing/OD/TypeA_OD1.log',mode="w"),
        logging.StreamHandler()
    ]
)

# Constants
MAX_WORKERS = 100
MAX_PROMPT_SIZE = 40000
SLICE_SIZE = 40000
MAX_RETRIES = 3
REQUEST_DELAY = 0.12
BATCH_SIZE = 500

# Google Sheet for token tracking
TOKEN_SHEET_ID = "1OvBOAXc_Y5aDLcK-BGCALFUZyJWLYolmFkr3tmo7mj4"
TOKEN_SHEET_NAME = "TypeA"


class Config:
    """Centralized configuration management"""
    SHEET_ID = "1eDlGi9iqnJ8gKmSN-y_QyAZzSNUyAkOvpBRnzO5Re88"
    PF_SHEET_NAME = "Funnels"
    FUNNEL_FILTER_IDS_SHEET_NAME = "Funnel filter ID's"
    EXTRACTING_SHEET_NAME = "DB"
    CREDENTIALS_FILE = "/home/tracxn-ds-528/Desktop/SR/Json_account/TypeA-1.json"
    API_BASE_URL = "https://platform.tracxn.com/api"
    GEMINI_API_URL = "https://generativelanguage.googleapis.com/v1beta/models/gemini-3.1-flash-lite:generateContent"

    HEADERS = {
        "accesstoken": "cf1ca0bf-3b85-49d9-9bc5-117d69cb1baa",
        "Content-Type": "application/json",
        "X-Request-Source": '123'
    }

    MAIN_HEADERS = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8',
        'Accept-Language': 'en-US,en;q=0.9',
    }


class GoogleSheetsClient:
    """Async Google Sheets client wrapper"""

    def __init__(self):
        self.client_manager = None
        self.scope = [
            'https://spreadsheets.google.com/feeds',
            'https://www.googleapis.com/auth/drive'
        ]

    async def initialize(self):
        """Initialize the client manager"""
        creds = ServiceAccountCredentials.from_json_keyfile_name(
            Config.CREDENTIALS_FILE,
            self.scope
        )
        self.client_manager = gspread_asyncio.AsyncioGspreadClientManager(lambda: creds)
        return self

    async def authorize(self):
        """Authorize the client"""
        if not self.client_manager:
            await self.initialize()
        return await self.client_manager.authorize()

    async def get_sheet_data(self, sheet_name: str, start_row: int = 2, start_col: int = 1) -> List[List[str]]:
        """Read data from a sheet"""
        gc = await self.authorize()
        sheet = await gc.open_by_key(Config.SHEET_ID)
        worksheet = await sheet.worksheet(sheet_name)
        data = await worksheet.get_all_values()
        return data[start_row - 1:]

    async def write_sheet_data(self, sheet_name: str, data: List[List[Any]], start_row: int,
                               start_col: int = 1) -> None:
        """Write data to a sheet"""
        gc = await self.authorize()
        sheet = await gc.open_by_key(Config.SHEET_ID)
        worksheet = await sheet.worksheet(sheet_name)

        end_col = start_col + len(data[0]) - 1
        end_row = start_row + len(data) - 1

        range_str = (
            f"{get_column_letter(start_col)}{start_row}:"
            f"{get_column_letter(end_col)}{end_row}"
        )

        await worksheet.update(range_str, data, value_input_option='USER_ENTERED')


class TracxnAPI:
    """Wrapper for Tracxn API operations"""

    def __init__(self, session: aiohttp.ClientSession):
        self.session = session
        self.headers = Config.HEADERS

    async def assign_funnel(self, domain_profile_id: str, funnel_id: str) -> Optional[str]:
        """Assign a funnel to a domain profile"""
        url = "https://platform.tracxn.com/data/funnel-action/force-assign"
        form_data = {
            "funnelId": funnel_id,
            "domainProfileId": domain_profile_id,
            "sourceDetails": {"source": "Write API"},
            "comment": "This is done by Write API"
        }

        async with self.session.put(url, json=form_data, headers=self.headers) as response:
            if response.status == 200:
                data = await response.json()
                return data.get("message", "No response message")
            logging.error(f"Error assigning funnel: {response.status}, {await response.text()}")
            return None

    async def move_out_funnel_sendback(self,domain_profile_id: str, funnel_id: str) -> str:
        """Move domain out of funnel"""
        
        print("sentbac selected")
        assign_status = await self.assign_funnel(domain_profile_id, funnel_id)
        if not assign_status:
            return "Assign Failed"
        url = "https://platform.tracxn.com/data/funnel-action/move"
        form_data = {
            "funnelId": funnel_id,
            "domainProfileId": domain_profile_id,
            "movedTo": ["64197f01a6dcff6572453ead"],
            "sourceDetails": {"source": "Write API"}
        }

        try:
            async with aiohttp.ClientSession() as session:
                async with session.put(url, json=form_data, headers=self.headers) as response:
                    if response.status in (200, 201,422):
                        await response.json()
                        return "Moved out to CF"
                    logging.error(f"Error moving funnel: {response.status}, {await response.text()}")
                    return "Error in Moving"
        except Exception as e:
            logging.error(f"Exception moving funnel: {e}")
            return "Error in Moving"


    async def move_out_funnel(self, domain_profile_id: str, funnel_id: str) -> str:
        """Move a domain profile out of a funnel"""
        url = "https://platform.tracxn.com/data/funnel-action/move"
        form_data = {
            "funnelId": funnel_id,
            "domainProfileId": domain_profile_id,
            "movedTo": ["5dc5863a2799a51cc0ff30e2"],
            "sourceDetails": {
                "source": "Write API",
                "sourceData": {
                    "view": "Card",
                    "tab": "Funnel Homepage"
                }
            }
        }

        async with self.session.put(url, json=form_data, headers=self.headers) as response:
            if response.status in (200, 201):
                return "Done"
            logging.error(f"Error moving funnel: {response.status}, {await response.text()}")
            return "NotDone"

    async def move_funnel_write_api(self, funnel_id: str, domain_profile_id: str) -> str:
        """Move funnel through write API"""
        
        await self.assign_funnel(domain_profile_id, funnel_id)
        await asyncio.sleep(0.5)
        return await self.move_out_funnel(domain_profile_id, funnel_id)

    async def update_company_sdld(self, domain_profile: str, sd: str, ld: str, tags: str) -> str:
        """Update company SD/LD data"""
        tags=tags.split(",")
        url = "https://platform.tracxn.com/data/entities/2.0/domain-profile"
        form_data = {
            "id": domain_profile,
            "description": {"value": ld},
            "shortDescription": {"value": sd},
            "keywords": {"value": {"HASHTAGS": tags}},
            # "publishingDepth": {"value": "Pub 1 - Full"}
            "publishingDepth": {"value": "Pub 2 - Partial"}

        }

        async with self.session.put(url, json=form_data, headers=self.headers) as response:
            if response.status in (200, 201):
                return "Done"
            logging.error(f"Error updating SDLD data for {domain_profile}: {response.status}, {await response.text()}")
            return "NotDone"

    async def update_company_sdldsf(self, domain_profile: str, sd: str, ld: str, tags: str,sf) -> str:
        """Update company SD/LD data"""
        tags=tags.split(",")
        tags.append("bu_Internal_SRprocess_TypeA")
        url = "https://platform.tracxn.com/data/entities/2.0/domain-profile"
        form_data = {
            "id": domain_profile,
            "description": {"value": ld},
            "shortDescription": {"value": sd},
            "keywords": {"value": {"HASHTAGS": tags}},
            # "publishingDepth": {"value": "Pub 1 - Full"},
            "publishingDepth": {"value": "Pub 2 - Partial"},
            "specialFlags": {
                "value": sf
            }
        }

        async with self.session.put(url, json=form_data, headers=self.headers) as response:
            if response.status in (200, 201):
                return "Done"
            logging.error(f"Error updating SDLD data for {domain_profile}: {response.status}, {await response.text()}")
            return "NotDone"

    async def update_company_bm(self, domain_profile: str, feedid: str, bmid: str) -> str:
        """Update company business model data"""
        url = "https://platform.tracxn.com/data/entities/3.0/w/theme-company-association"

        form_data = {
           "object": {
             "themeId": feedid,
             "status": "PUBLISHED",
             "businessModelId": bmid,
             "companyId": domain_profile,
            },
            "opType":"Update"

        }

        async with self.session.put(url, json=form_data, headers=self.headers) as response:
            if response.status in (200, 201):
                return "Done"
            logging.error(f"Error updating BM data for {domain_profile}: {response.status}, {await response.text()}")
            return response.status

    async def remove_hashtag(self, tags: str, domain_profile: str) -> str:
        """Remove hashtag from company"""
        tags_list = [tag for tag in tags.split(",")
                     if tag != "llmbasedpublishing" and tag != "bu_llm_sd_ld"]
        form_data = {
            "id": domain_profile,
            "keywords": {"value": {"HASHTAGS": tags_list}}
        }

        async with self.session.put(
                "https://platform.tracxn.com/data/entities/2.0/domain-profile",
                json=form_data,
                headers=self.headers
        ) as response:
            if response.status in (200, 201):
                return "Done"
            logging.error(f"Error removing hashtag for {domain_profile}: {response.status}, {await response.text()}")
            return "NotDone"


class WebScraper:
    """Handles web scraping operations"""

    def __init__(self, session: aiohttp.ClientSession):
        self.session = session

    async def fetch_page(self, url: str, retries: int = MAX_RETRIES, delay: int = 2) -> Tuple[Optional[str], Any]:
        """Fetch a web page with retries and fallback protocols"""
        for attempt in range(retries):
            try:
                logging.info(f"Fetching URL: {url} (Attempt {attempt + 1}/{retries})")

                # Try different protocols on subsequent attempts
                if attempt == 1:
                    url = url.replace("https://", "http://www.")
                elif attempt == 2:
                    url = url.replace("http://www.", "https://www.")

                async with self.session.get(
                        url,
                        timeout=10,
                        allow_redirects=True,
                        headers=Config.MAIN_HEADERS
                ) as response:
                    logging.info(f"Response status: {response.status} for URL: {url}")
                    raw_bytes = await response.read()
                    encoding = response.get_encoding() or "utf-8"

                    try:
                        return raw_bytes.decode(encoding), response.status
                    except UnicodeDecodeError:
                        return raw_bytes.decode("iso-8859-1"), response.status
            except (aiohttp.ClientError, asyncio.TimeoutError) as e:
                logging.warning(f"Error fetching {url}: {e}")
                if attempt < retries - 1:
                    await asyncio.sleep(delay)

        return None, "Max retries exceeded"

    @staticmethod
    def clean_html(html: str) -> str:
        """Clean HTML content by removing scripts, styles, and tags"""
        if not html:
            return ""

        html = re.sub(r'<script[^>]*>[\s\S]*?</script>', '', html, flags=re.IGNORECASE)
        html = re.sub(r'<style[^>]*>[\s\S]*?</style>', '', html, flags=re.IGNORECASE)
        html = re.sub(r'<[^>]+>', '', html)
        html = re.sub(r'\s+', ' ', html).strip()
        return html

    @staticmethod
    def extract_links(html: str, domain: str) -> Set[str]:
        """Extract all links from HTML content"""
        soup = BeautifulSoup(html, "html.parser")
        links = set()

        for tag in soup.find_all("a", href=True):
            href = tag["href"].strip()

            if href.startswith("/"):
                href = f"https://{domain}{href}"
            elif href.startswith(("http://", "https://")):
                links.add(href)

        return links

    async def scrape_with_bs(self, url: str) -> Tuple[Optional[str], Any]:
        """Scrape a webpage using BeautifulSoup"""
        html, status_code = await self.fetch_page(url)
        return self.clean_html(html) if html else None, status_code


class GeminiAPI:
    """Handles interactions with Gemini API"""
    gemini_calls = 0
    total_input_tokens = 0
    total_output_tokens = 0

    def __init__(self, api_key: str):
        self.api_key = api_key

    async def call_gemini_api(self, prompt: str) -> str:
        """Call Gemini API with the given prompt"""
        if not prompt:
            return ""

        GeminiAPI.gemini_calls += 1
        headers = {"Content-Type": "application/json"}
        data = {
            "contents": [{"parts": [{"text": prompt}]}],
            "generationConfig": {"temperature": 0},
        }

        url = f"{Config.GEMINI_API_URL}?key={self.api_key}"

        async with aiohttp.ClientSession() as session:
            try:
                async with session.post(url, headers=headers, json=data) as response:
                    response_json = await response.json()

                    # ?? Token tracking ??????????????????????????????????????
                    usage = response_json.get("usageMetadata", {})
                    input_tokens  = usage.get("promptTokenCount", 0)
                    output_tokens = usage.get("candidatesTokenCount", 0)
                    GeminiAPI.total_input_tokens  += input_tokens
                    GeminiAPI.total_output_tokens += output_tokens
                    logging.info(
                        f"Gemini call #{GeminiAPI.gemini_calls} ? "
                        f"in: {input_tokens}, out: {output_tokens} | "
                        f"running total ? in: {GeminiAPI.total_input_tokens}, "
                        f"out: {GeminiAPI.total_output_tokens}"
                    )
                    # ????????????????????????????????????????????????????????

                    return (
                        response_json
                        .get("candidates", [{}])[0]
                        .get("content", {})
                        .get("parts", [{}])[0]
                        .get("text", "")
                    )
            except Exception as e:
                logging.error(f"Error calling Gemini API: {e}")
                return ""

    @staticmethod
    def extract_descriptions(text: str) -> Tuple[str, str]:
        # First try to extract from JSON-like format
        json_sd_match = re.search(r'["\']Short Description["\']:\s*["\'](.*?)["\']', text, re.DOTALL)
        json_ld_match = re.search(r'["\']Long Description["\']:\s*["\'](.*?)["\']', text, re.DOTALL)

        if json_sd_match and json_ld_match:
            short_desc = json_sd_match.group(1).strip()
            long_desc = json_ld_match.group(1).strip()
        else:
            # Fall back to regular text parsing
            sd_match = re.search(r"Short Description:\s*(.*?)(?=\nLong Description:|\n\n|$)", text, re.DOTALL)
            ld_match = re.search(r"Long Description:\s*(.*)", text, re.DOTALL)
            short_desc = sd_match.group(1).strip() if sd_match else ""
            long_desc = ld_match.group(1).strip() if ld_match else ""

        # Clean up whitespace (replace multiple spaces/newlines with single space)
        short_desc = " ".join(short_desc.split())
        long_desc = " ".join(long_desc.split())

        # Remove trailing period from short description if present
        if short_desc.endswith('.'):
            short_desc = short_desc[:-1].strip()

        return short_desc, long_desc


class DataProcessor:
    """Processes and manages data operations"""

    def __init__(self, client: GoogleSheetsClient):
        self.client = client
        self.main_prompts = None
        self.match_paths = None
        self.feed_id_map = None
        self.bm_mapping = None
        self.bm_ids = None
        self.feed_def_map = None
        self.bm_1stlvl_live_stat=None

    async def initialize_data(self):
        """Initialize all required data mappings"""
        await self.get_main_prompts()
        await self.get_match_paths()
        await self.get_feed_id_map()
        await self.get_bm_mapping()
        await self.get_feed_def_map()

    async def get_main_prompts(self) -> List[str]:
        """Get main prompts from Google Sheet"""
        if self.main_prompts is not None:
            return self.main_prompts

        gc = await self.client.authorize()
        sheet = await gc.open_by_key("1N9GgEXIiR7QwEpzpJCvGbXlZ_kCgN9ev8fj0Ynv98MU")
        worksheet = await sheet.worksheet("Prompts")
        data = await worksheet.get_all_values()
        self.main_prompts = [row[1] for row in data[1:10]]
        return self.main_prompts

    async def get_match_paths(self) -> List[List[str]]:
        """Get match paths from Google Sheet"""
        if self.match_paths is not None:
            return self.match_paths

        gc = await self.client.authorize()
        sheet = await gc.open_by_key(Config.SHEET_ID)
        worksheet = await sheet.worksheet("Paths")
        data = await worksheet.get_all_values()
        self.match_paths = [
            [cell for cell in row if cell.strip()]
            for row in data if any(cell.strip() for cell in row)
        ]
        return self.match_paths

    async def get_feed_id_map(self) -> Dict[str, str]:
        """Get feed ID mapping from Google Sheet"""
        if self.feed_id_map is not None:
            return self.feed_id_map

        gc = await self.client.authorize()
        sheet = await gc.open_by_key("1VSvvKsjO5ZPSg3ff6SnwPEQ0i9BTwzAI-aCiWxjHzYU")
        worksheet = await sheet.worksheet("Feed Owner Details")
        data = await worksheet.get_all_values()
        self.feed_id_map = {row[0]: row[1] for row in data if len(row) > 1}
        return self.feed_id_map

    async def get_bm_mapping(self) -> Tuple[Dict[str, Any], Dict[str, str]]:
        """Get business model mapping from Google Sheets"""
        if self.bm_mapping is not None and self.bm_ids is not None:
            return self.bm_mapping, self.bm_ids,self.bm_1stlvl_live_stat

        try:
            gc = await self.client.authorize()
            sheet = await gc.open_by_key("1hi_Zb_0DsK8CRqWST3_FrrjG3tj5NoMKrZHUroYLcOw")

            # Initialize mappings
            self.bm_mapping = {}
            self.bm_ids = {}
            self.bm_1stlvl_live_stat={}

            # Process 1st Level Sheet
            first_level_ws = await sheet.worksheet("1st Level")
            first_level_data = await first_level_ws.get_all_values()

            for row in first_level_data[1:]:
                if len(row) < 6:
                    continue

                feed = row[0]
                bm_path = row[1]
                bm_id = row[2]
                stat = row[5]
                description = row[4] if row[4] != "-" else row[3]

                if not feed or not bm_path:
                    continue
                self.bm_ids[bm_path] = bm_id
                self.bm_1stlvl_live_stat[bm_path] = stat

                if feed not in self.bm_mapping:
                    self.bm_mapping[feed] = {
                        "1stLevel": [],
                        "2ndLevel": [],
                        "1stLevelCount": 1,
                        "2ndLevelCount": 1
                    }

                self.bm_mapping[feed]["1stLevel"].append([
                    self.bm_mapping[feed]["1stLevelCount"],
                    bm_path,
                    description
                ])
                self.bm_mapping[feed]["1stLevelCount"] += 1

            # Process 2nd Level Sheet
            second_level_ws = await sheet.worksheet("2nd Level Live BM's")
            second_level_data = await second_level_ws.get_all_values()

            for row in second_level_data[1:]:
                if len(row) < 4:
                    continue

                feed = row[0]
                bm_path = row[1]
                bm_id = row[2]
                description = row[3]

                if not feed or not bm_path:
                    continue

                self.bm_ids[bm_path] = bm_id

                if feed not in self.bm_mapping:
                    self.bm_mapping[feed] = {
                        "1stLevel": [],
                        "2ndLevel": [],
                        "1stLevelCount": 1,
                        "2ndLevelCount": 1
                    }

                self.bm_mapping[feed]["2ndLevel"].append([
                    self.bm_mapping[feed]["2ndLevelCount"],".",
                     bm_path,
                    " -"+description
                ])
                self.bm_mapping[feed]["2ndLevelCount"] += 1

        except Exception as e:
            logging.error(f"Error fetching BM mapping data: {e}")

        return self.bm_mapping, self.bm_ids,self.bm_1stlvl_live_stat

    async def get_feed_def_map(self) -> Dict[str, str]:
        """Get feed definition mapping"""
        if self.feed_def_map is not None:
            return self.feed_def_map

        gc = await self.client.authorize()
        sheet = await gc.open_by_key("1HEmWY4AeFltmjPbMzX-xDydTsncMX53hpbgHpsS_-44")
        worksheet = await sheet.worksheet("Feed Definition")
        data = await worksheet.get_all_values()
        self.feed_def_map = {row[1]: row[3] for row in data if len(row) > 1}
        return self.feed_def_map


class DomainProcessor:
    """Processes individual domains"""

    def __init__(
            self,
            session: aiohttp.ClientSession,
            client: GoogleSheetsClient,
            data_processor: DataProcessor,
            gemini: GeminiAPI,
            tracxn_api: TracxnAPI,
            scraper: WebScraper,
            semaphore: asyncio.Semaphore
    ):
        self.session = session
        self.client = client
        self.data_processor = data_processor
        self.gemini = gemini
        self.tracxn_api = tracxn_api
        self.scraper = scraper
        self.semaphore = semaphore

    async def process_domain(
            self,
            domain: str,
            row: int,
            funnel_name: str,
            funnel_id: str,
            domain_profile_id: str,
            hashtags: str,
            PFSF: List[any],
            feed_def_map
    ) -> List[Any]:
        """Process a single domain with all its operations"""
        async with self.semaphore:
            try:
                main_prompts = await self.data_processor.get_main_prompts()
                match_paths = await self.data_processor.get_match_paths()
                bm_mapping, bm_ids,bm_1stlvel_check = await self.data_processor.get_bm_mapping()
                feed_id_map = await self.data_processor.get_feed_id_map()
                feed_def_map = await self.data_processor.get_feed_def_map()

                # Scrape the domain
                url = f"https://{domain}"
                home_body, status_code = await self.scraper.scrape_with_bs(url)
                if not home_body:
                    return await self.handle_no_data_case(
                        row,
                        domain_profile_id,
                        hashtags,
                        funnel_id
                    )

                # Extract links from homepage
                html, _ = await self.scraper.fetch_page(url)
                links = self.scraper.extract_links(html, domain) if html else set()

                # Process additional pages
                body_results, extracted_pages = await self.process_additional_pages(
                    home_body,
                    links,
                    match_paths,
                    url
                )

                # Generate prompts
                prompt1, prompt2 = self.generate_prompts(body_results)

                # Process Gemini responses
                sd_p1, ld_p1, sd_p2, ld_p2 = await self.process_gemini_responses(
                    prompt1,
                    prompt2,
                    main_prompts
                )

                # Combine descriptions
                ld_main = f"{ld_p1}\n\n{ld_p2}" if ld_p1 or ld_p2 else ""
                # Handle case where we couldn't get enough data
                if not sd_p1 or not ld_main or not ld_p1:
                    return await self.handle_no_data_case(
                        row,
                        domain_profile_id,
                        hashtags,
                        funnel_id
                    )

                # Process business model data
                feed = funnel_name.split(" : ")[1]
                feed_id = feed_id_map.get(feed, "")
                feed_definition = feed_def_map.get(feed, "")
                return await self.process_business_model(
                    row,
                    domain,
                    domain_profile_id,
                    hashtags,
                    feed,
                    feed_id,
                    ld_main,
                    bm_mapping,
                    bm_ids,
                    funnel_id,
                    sd_p1,
                    ld_p1,
                    ld_p2,
                    main_prompts,
                    PFSF,
                    feed_definition,
                    bm_1stlvel_check
                )

            except Exception as e:
                logging.error(f"Error processing domain {domain}: {e}")
                return []

    async def handle_no_data_case(
            self,
            row: int,
            domain_profile_id: str,
            hashtags: str,
            funnel_id: str
    ) -> List[Any]:
        """Handle case where we couldn't get web data"""
        hash_status = await self.tracxn_api.remove_hashtag(hashtags, domain_profile_id)
        funnel_update = await self.tracxn_api.move_out_funnel_sendback(domain_profile_id,funnel_id)
        # hash_status="Need to remove hash"
        sheet_data = [
            "No",
            "", "", "", "", "", "", "", "", "",
            hash_status,
            "",
            "NoWebScrap",
            "NoWebScrap",
            funnel_update,
            ""
        ]

        await self.client.write_sheet_data(
            Config.EXTRACTING_SHEET_NAME,
            [sheet_data],
            row,
            8
        )
        return sheet_data

    async def process_additional_pages(
            self,
            home_body: str,
            links: Set[str],
            match_paths: List[List[str]],
            url: str
    ) -> Tuple[List[str], List[str]]:
        """Process additional pages based on match paths"""
        body_results = [home_body]
        extracted_pages = ["homepage"]

        for i, path_group in enumerate(match_paths):
            extracted_data = "nodata"

            for path in path_group:
                matched_link = next(
                    (link for link in links if path in link),
                    None
                )

                if matched_link:
                    page_html, page_status = await self.scraper.fetch_page(matched_link)

                    if page_html and page_status == 200:
                        extracted_data = self.scraper.clean_html(page_html)
                        extracted_pages.append([
                                                   "About Page", "Product Page", "Pricing Page",
                                                   "Solutions Page", "Offerings Page", "Features Page",
                                                   "Technology Page", "Mission Page",
                                                   "Impact/Social Impact Page", "Team Page"
                                               ][i])
                        break

            body_results.append(extracted_data)

        return body_results, extracted_pages

    def generate_prompts(self, body_results: List[str]) -> Tuple[str, str]:
        """Generate prompts from body results"""
        prompt1 = ""
        if body_results[0] != "nodata":
            prompt1 += body_results[0]
        if body_results[1] != "nodata":
            prompt1 += "\n\n" + body_results[1]

        prompt1 = prompt1[:SLICE_SIZE] if len(prompt1) > MAX_PROMPT_SIZE else prompt1

        prompt2 = ""
        for body in body_results[2:]:
            if body != "nodata":
                prompt2 += "\n\n" + body

        prompt2 = prompt2[:SLICE_SIZE] if len(prompt2) > MAX_PROMPT_SIZE else prompt2

        return prompt1, prompt2

    async def process_gemini_responses(
            self,
            prompt1: str,
            prompt2: str,
            main_prompts: List[str]
    ) -> Tuple[str, str, str, str]:
        """Process Gemini API responses"""
        sd_p1, ld_p1 = "", ""
        sd_p2, ld_p2 = "", ""

        if len(prompt1) > 100:
            final_prompt1 = main_prompts[0].replace("XX", prompt1)
            gemini_response_p1 = await self.gemini.call_gemini_api(final_prompt1)
            sd_p1, ld_p1 = self.gemini.extract_descriptions(gemini_response_p1)

        if len(prompt2) > 100:
            final_prompt2 = main_prompts[1].replace("XX", prompt2)
            gemini_response_p2 = await self.gemini.call_gemini_api(final_prompt2)
            sd_p2, ld_p2 = self.gemini.extract_descriptions(gemini_response_p2)

        return sd_p1, ld_p1, sd_p2, ld_p2

    async def process_business_model(
            self,
            row: int,
            domain: str,
            domain_profile_id: str,
            hashtags: str,
            feed: str,
            feed_id: str,
            ld_main: str,
            bm_mapping: Dict[str, Any],
            bm_ids: Dict[str, str],
            funnel_id: str,
            sd_p1: str,
            ld_p1: str,
            ld_p2: str,
            main_prompts: List[str],
            PFSF: List[any],
            feed_definition,
            bm_1stlevel_check
    ) -> List[Any]:
        """Process business model data"""
        # Initialize all variables with default values to avoid UnboundLocalError
        company_edits = "NotUpdated"
        bm_update = "NotUpdated"
        funnel_update = "NotUpdated"
        bm_name_1stlevel = "No Results"
        bm_name_2ndlevel = "No BM matched"
        bm_id_2ndlevel = "No ID"
        gemini_response_bm_1stlevel = ""
        gemini_response_bm_2ndlevel = ""
        bm_prompt_1stlevel = ""
        bm_prompt_2ndlevel = ""
        output = ""
        sfarray = []
        SFresult = {'YES': [], 'NO': []}  # Initialize SFresult here to avoid access errors

        try:
            hashtags_list = hashtags.split(",")
            hashtags_list.append("bu_llm_sd_ld")

            # Process 1st level BM
            bm_prompt_1stlevel = main_prompts[6].replace("YY", ld_main)
            bm_prompt_1stlevel = bm_prompt_1stlevel.replace("XX", feed_definition)
            feed_bms_1stlevel = bm_mapping.get(feed, {}).get("1stLevel", [])

            if not feed_bms_1stlevel:
                company_edits = await self.tracxn_api.update_company_sdld(
                    domain_profile_id,
                    sd_p1,
                    ld_p1,
                    ','.join(hashtags_list)
                )
                funnel_update = await self.tracxn_api.move_out_funnel_sendback(domain_profile_id,funnel_id)
                return await self.handle_no_bm_case(
                    row,
                    domain_profile_id,
                    ','.join(hashtags_list),
                    feed_id,
                    "No BMs for this feed",
                    company_edits,
                    sd_p1,
                    ld_p1,
                    ld_p2,
                    bm_prompt_1stlevel,
                    "No BMs available",
                    funnel_update
                )

            id_to_category_map_1stlevel = {row[0]: row[1] for row in feed_bms_1stlevel}
            feed_1stlevel_bms_str = "\n".join(
                [f"{row[0]}. {row[1]} - {row[2]}" for row in feed_bms_1stlevel]
            )

            bm_prompt_1stlevel = bm_prompt_1stlevel.replace("BM_Paths", feed_1stlevel_bms_str)
            print("len of bmprompt 1st lvel ")
            print(len(bm_prompt_1stlevel))
            gemini_response_bm_1stlevel = await self.gemini.call_gemini_api(bm_prompt_1stlevel)

            # Extract BM from response
            try:
                if (isinstance(gemini_response_bm_1stlevel, str) and
                        gemini_response_bm_1stlevel.strip().startswith(("No Results", "No results"))):
                    bm_name_1stlevel = "No Results"
                else:
                    pattern = pattern = r'^\d+[\.\s]+\s*(.*?)\s*[,:-]\s*Explanation'
                    match = re.search(pattern, gemini_response_bm_1stlevel, re.MULTILINE)
                    if match:
                        bm_name_1stlevel = match.group(1).strip()
                        bm_id = bm_ids.get(bm_name_1stlevel, "")
                    else:
                        bm_name_1stlevel = "No Results"
            except Exception:
                bm_name_1stlevel = "No Results"

            if bm_name_1stlevel == "No Results":
                company_edits = await self.tracxn_api.update_company_sdld(
                    domain_profile_id,
                    sd_p1,
                    ld_p1,
                    ','.join(hashtags_list)
                )
                funnel_update = await self.tracxn_api.move_out_funnel_sendback(domain_profile_id,funnel_id)
                return await self.handle_no_bm_case(
                    row,
                    domain_profile_id,
                    ','.join(hashtags_list),
                    feed_id,
                    "No BM matched",
                    company_edits,
                    sd_p1,
                    ld_p1,
                    ld_p2,
                    bm_prompt_1stlevel,
                    gemini_response_bm_1stlevel,
                    funnel_update
                )

            # Process 2nd level BM if available
            feed_bms_2ndlevel = bm_mapping.get(feed, {}).get("2ndLevel", [])
           
            if feed_bms_2ndlevel:
                bm_prompt_2ndlevel = main_prompts[7].replace("XX", ld_main)

                filtered_array = [
                    subarray for subarray in feed_bms_2ndlevel
                    if subarray[2].startswith(bm_name_1stlevel)
                ]
               
                if filtered_array:
                    feed_2ndlevel_bms_str = "\n".join([" ".join(map(str, row)) for row in filtered_array])
                    bm_prompt_2ndlevel = bm_prompt_2ndlevel.replace("BM_Paths", feed_2ndlevel_bms_str)
                    print("len of bmprompt 2nd lvel ")
                    print(len(bm_prompt_2ndlevel))
                    gemini_response_bm_2ndlevel = await self.gemini.call_gemini_api(bm_prompt_2ndlevel)

                    try:
                        if (isinstance(gemini_response_bm_2ndlevel, str) and
                                gemini_response_bm_2ndlevel.strip().startswith(("No Results", "No results"))):
                            if bm_1stlevel_check.get(bm_name_1stlevel) == "Live":
                                bm_name_2ndlevel = bm_name_1stlevel
                                bm_id_2ndlevel = bm_ids.get(bm_name_1stlevel, "no id")
                            else:
                                bm_name_2ndlevel = "No BM matched"
                                bm_id_2ndlevel = "no id"
                        else:
                            pattern = r'^\d+[\.\s]+\s*(.*?)\s*[,:-]\s*Explanation'
                            match = re.search(pattern, gemini_response_bm_2ndlevel, re.MULTILINE)
                            if match:
                                bm_name_2ndlevel = match.group(1).strip()
                                bm_id_2ndlevel = bm_ids.get(bm_name_2ndlevel, "no id")
                            else:
                                if bm_1stlevel_check.get(bm_name_1stlevel) == "Live":
                                    bm_name_2ndlevel = bm_name_1stlevel
                                    bm_id_2ndlevel = bm_ids.get(bm_name_1stlevel, "no id")
                                else:
                                    bm_name_2ndlevel = "No BM matched"
                                    bm_id_2ndlevel = "no id"
                    except Exception:
                        if bm_1stlevel_check.get(bm_name_1stlevel) == "Live":
                            bm_name_2ndlevel = bm_name_1stlevel
                            bm_id_2ndlevel = bm_ids.get(bm_name_1stlevel, "no id")
                        else:
                            bm_name_2ndlevel = "No BM matched"
                            bm_id_2ndlevel = "no id"

            bm_prompt_2ndlevel = bm_prompt_2ndlevel[:SLICE_SIZE] if len(bm_prompt_2ndlevel) > MAX_PROMPT_SIZE else bm_prompt_2ndlevel
            # Process special flags
            # hashtags_list.append("bu_llm_typeA_autopublish")

            if bm_id_2ndlevel != "no id" and bm_id_2ndlevel!=None and bm_id_2ndlevel!="No ID":
                sfprompt = main_prompts[8].replace("XX", ld_p1 + ld_p2)
                flags = await self.gemini.call_gemini_api(sfprompt)
                sfarray = self.merge_and_extract_json(flags, json.loads(PFSF) if PFSF else [])

                id_to_category = {
                    1: "Consumer", 2: "Enterprise", 6: "Marketplace", 7: "SaaS",
                    9: "Tech", 11: "Software", 12: "Tech Hardware", 13: "Social Impact",
                    14: "Innovative", 15: "Artificial Intelligence", 16: "High IP",
                    17: "Goods", 18: "Service", 19: "Looks Scalable", 20: "Chain",
                    21: "Not for Profit", 22: "Blockchain", 23: "Native AI"
                }

                # Reset SFresult to empty before populating
                SFresult = {'YES': [], 'NO': []}

                # Update company with special flags
                company_edits = await self.tracxn_api.update_company_sdldsf(
                    domain_profile_id,
                    sd_p1,
                    ld_p1,
                    ','.join(hashtags_list),
                    sfarray
                )

            else:
                company_edits=await self.tracxn_api.update_company_sdld(
                    domain_profile_id,
                    sd_p1,
                    ld_p1,
                    ','.join(hashtags_list)
                )
                funnel_update = await self.tracxn_api.move_out_funnel_sendback(domain_profile_id,funnel_id)
                nobmmatcheddata=[
                "Yes",
                sd_p1,
                ld_p1,
                ld_p2,
                bm_prompt_1stlevel,
                f"{bm_name_1stlevel}\n\n{gemini_response_bm_1stlevel}",
                bm_prompt_2ndlevel if feed_bms_2ndlevel else "",
                bm_name_2ndlevel,
                gemini_response_bm_2ndlevel if feed_bms_2ndlevel else "",
                bm_id_2ndlevel,
                "No need to remove",
                feed_id,
                company_edits,
                bm_update,
                funnel_update,
                output
                    ]
               
                await self.client.write_sheet_data(
                Config.EXTRACTING_SHEET_NAME,
                [nobmmatcheddata],
                row,
                8
                    )  
                return nobmmatcheddata


            # Update business model
            if bm_id_2ndlevel != "no id" and company_edits=="Done":
                bm_update = await self.tracxn_api.update_company_bm(
                    domain_profile_id,
                    feed_id,
                    bm_id_2ndlevel
                )

            # Move funnel if updates were successful
            print(bm_update)
            if company_edits == "Done" and (bm_update == "Done" or bm_update == 422):
                funnel_update = await self.tracxn_api.move_funnel_write_api(funnel_id, domain_profile_id)

            # Format special flags output - only if sfarray was populated
            if sfarray:
                for item in sfarray:
                    category = id_to_category.get(item['id'], f"Unknown ID {item['id']}")
                    SFresult[item['value']].append(category)

                output = f"YES : {','.join(SFresult['YES'])}\nNO : {','.join(SFresult['NO'])}"
            else:
                output = "No special flags processed"

            # Prepare final data
            if(output == "No special flags processed"):
                bm_update="Error"

            sheet_data = [
                "Yes",
                sd_p1,
                ld_p1,
                ld_p2,
                bm_prompt_1stlevel,
                f"{bm_name_1stlevel}\n\n{gemini_response_bm_1stlevel}",
                bm_prompt_2ndlevel if feed_bms_2ndlevel else "",
                bm_name_2ndlevel,
                gemini_response_bm_2ndlevel if feed_bms_2ndlevel else "",
                bm_id_2ndlevel,
                "No need to remove",
                feed_id,
                company_edits,
                bm_update,
                funnel_update,
                output
            ]

            await self.client.write_sheet_data(
                Config.EXTRACTING_SHEET_NAME,
                [sheet_data],
                row,
                8
            )

            return sheet_data

        except Exception as e:
            logging.error(f"Error processing business model for {domain}: {e}")
            # Return error data
            error_data = [
                "Error",
                sd_p1,
                ld_p1,
                ld_p2,
                bm_prompt_1stlevel,
                str(e),
                "", "", "", "",
                "Error",
                feed_id,
                "Error",
                "Error",
                "Error",
                ""
            ]

            await self.client.write_sheet_data(
                Config.EXTRACTING_SHEET_NAME,
                [error_data],
                row,
                8
            )

            return error_data

    async def handle_no_bm_case(
            self,
            row: int,
            domain_profile_id: str,
            hashtags: str,
            feed_id: str,
            error_msg: str,
            company_edits: str,
            sd_p1 : str,
            ld_p1: str,
            ld_p2:str,
            bm_prompt_1stlevel:str,
            gemini_response_bm_1stlevel:str,
            funnel_update:str
    ) -> List[Any]:
        """Handle case where no BM could be determined"""
        hash_status = await self.tracxn_api.remove_hashtag(hashtags, domain_profile_id)

        sheet_data = [
            "Yes",
            sd_p1, ld_p1, ld_p2, bm_prompt_1stlevel, gemini_response_bm_1stlevel, "", "", "", "",
            hash_status,
            feed_id,
            company_edits,
            error_msg,
            funnel_update,
            ""
        ]

        await self.client.write_sheet_data(
            Config.EXTRACTING_SHEET_NAME,
            [sheet_data],
            row,
            8
        )

        return sheet_data

    @staticmethod
    def merge_and_extract_json(input_text, sf_arr=None):
        """
        Improved version with better error handling
        """
        # Handle case where input is already an array
        if isinstance(input_text, list):
            afs_arr = input_text
        else:
            # Try multiple JSON extraction patterns
            json_patterns = [
                r'(\[\s*\{.*?\}\s*\])',  # Original pattern
                r'```json\s*(\[\s*\{.*?\}\s*\])\s*```',  # Code block pattern
                r'JSON:\s*(\[\s*\{.*?\}\s*\])'  # Labeled JSON
            ]
           
            json_str = None
            for pattern in json_patterns:
                match = re.search(pattern, str(input_text), re.DOTALL)
                if match:
                    json_str = match.group(1)
                    break
           
            if not json_str:
                logging.error("No JSON array found in input text")
                return None
               
            try:
                # Clean the JSON string
                json_str = json_str.replace("'", '"')  # Single to double quotes
                json_str = re.sub(r':\s*(YES|NO)\s*([,}])', r': "\1"\2', json_str)  # Quote YES/NO
                json_str = re.sub(r',\s*\]', ']', json_str)  # Remove trailing commas
               
                afs_arr = json.loads(json_str)
            except json.JSONDecodeError as e:
                logging.error(f"JSON decode failed: {e}")
                logging.error(f"Cleaned JSON: {json_str}")
                return None

        # Rest of your code...
        sf_arr = sf_arr or []
       
        try:
            afs_dict = {item['id']: item['value'] for item in afs_arr if 'id' in item}
            sf_dict = {item['id']: item['value'] for item in sf_arr if 'id' in item}
            merged_dict = {**afs_dict, **sf_dict}
            return [{'id': k, 'value': v} for k, v in merged_dict.items()]
        except (KeyError, TypeError) as e:
            logging.error(f"Error merging arrays: {e}")
            return None


async def feed_def(client: GoogleSheetsClient) -> Dict[str, str]:
    """Get feed definitions with caching"""
    gc = await client.authorize()

    # First sheet
    sheet1 = await gc.open_by_key("1BhWtCW8j3ixsXf8_uO8GjaMYl_txhAB-561Hbyy9jh4")
    worksheet1 = await sheet1.worksheet("Feed Definition (Worked)")
    data1 = await worksheet1.get_all_values()
    data_dict = {row[1]: row[4] for row in data1 if len(row) > 4}

    # Second sheet
    sheet2 = await gc.open_by_key("1zL2hMn6FhSnuAw0Qu7RHj3oAEFxchzBJZKLNg5Y4mM4")
    worksheet2 = await sheet2.worksheet("Feed Definition (Worked)")
    data2 = await worksheet2.get_all_values()

    # Merge data
    for row in data2:
        if len(row) > 4 and row[1] not in data_dict:
            data_dict[row[1]] = row[4]

    return data_dict


async def flush_token_counts(client: GoogleSheetsClient) -> None:
    """
    Read the current input/output token counts from the tracking sheet,
    add this run's totals, and write the updated values back.

    Sheet: TOKEN_SHEET_ID / TOKEN_SHEET_NAME
      B2 = cumulative input  tokens
      B3 = cumulative output tokens
    """
    try:
        gc = await client.authorize()
        sheet = await gc.open_by_key(TOKEN_SHEET_ID)
        worksheet = await sheet.worksheet(TOKEN_SHEET_NAME)

        # Read existing counts (B2 and B3)
        b2_val = await worksheet.acell("B2")
        b3_val = await worksheet.acell("B3")

        existing_input  = int(b2_val.value or 0)
        existing_output = int(b3_val.value or 0)

        new_input  = existing_input  + GeminiAPI.total_input_tokens
        new_output = existing_output + GeminiAPI.total_output_tokens

        # Write incremented values back
        await worksheet.update("B2:B3", [[new_input], [new_output]], value_input_option="USER_ENTERED")

        logging.info(
            f"Token counts updated ? "
            f"input: {existing_input} + {GeminiAPI.total_input_tokens} = {new_input} | "
            f"output: {existing_output} + {GeminiAPI.total_output_tokens} = {new_output}"
        )
    except Exception as e:
        logging.error(f"Failed to flush token counts to sheet: {e}")


async def main():
    """Main async function to run the scraper"""
    try:
        # Initialize clients and services
        client = await GoogleSheetsClient().initialize()
        data_processor = DataProcessor(client)

        # Read existing data
        pf_data = await client.get_sheet_data(Config.EXTRACTING_SHEET_NAME)
        feed_def_map = await feed_def(client)
        # Initialize semaphore for concurrency control
        semaphore = asyncio.Semaphore(MAX_WORKERS)

        # Process domains in batches
        async with aiohttp.ClientSession(headers=Config.HEADERS) as session:
            # Initialize services
            tracxn_api = TracxnAPI(session)
            scraper = WebScraper(session)
            gemini = GeminiAPI("AIzaSyBB072_z22KVzuW-JGu8tGL-WVoxxlfuRM")

            # Initialize data processor
            await data_processor.initialize_data()

            # Process domains in batches
            for i in range(0, len(pf_data), BATCH_SIZE):
                batch = pf_data[i:i + BATCH_SIZE]
                tasks = []

                for idx, row in enumerate(batch, start=i + 2):
                    if not row or len(row) < 6 or row[7]:  # Skip processed rows
                        continue

                    try:
                        domain = row[1]
                        funnel_name = row[3]
                        funnel_id = row[4]
                        domain_profile_id = row[2]
                        hashtags = row[5]
                        PFSF=row[6]

                        # Create domain processor
                        domain_processor = DomainProcessor(
                            session,
                            client,
                            data_processor,
                            gemini,
                            tracxn_api,
                            scraper,
                            semaphore

                        )

                        # Create task
                        task = asyncio.create_task(
                            domain_processor.process_domain(
                                domain,
                                idx,
                                funnel_name,
                                funnel_id,
                                domain_profile_id,
                                hashtags,
                                PFSF,
                                feed_def_map
                            )
                        )

                        tasks.append(task)
                        await asyncio.sleep(REQUEST_DELAY)

                    except ValueError as e:
                        logging.error(f"Error unpacking row: {row}, Error: {e}")
                        continue

                # Wait for current batch to complete
                await asyncio.gather(*tasks)
                logging.info(f"Completed processing batch of {len(batch)} domains.")

    except Exception as e:
        logging.error(f"Fatal error in main: {e}", exc_info=True)
    finally:
        # ?? Flush accumulated token counts to the tracking sheet ??
        try:
            await flush_token_counts(client)
        except Exception as e:
            logging.error(f"Could not flush token counts: {e}")
        logging.info("Script execution completed.")


if __name__ == "__main__":
    asyncio.run(main())